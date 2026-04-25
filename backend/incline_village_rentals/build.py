"""
Build the Incline Village rental-gear SQLite database AND the
companion Excel spreadsheet from `data.py`.

Run from the backend directory:

    python -m incline_village_rentals.build

Outputs (alongside this file):
    incline_village_rentals.db            — SQLite database
    Incline_Village_Rental_Gear.xlsx      — formatted spreadsheet
"""

from __future__ import annotations

import os
import sqlite3
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .data import BUSINESSES

HERE = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(HERE, "incline_village_rentals.db")
XLSX_PATH = os.path.join(HERE, "Incline_Village_Rental_Gear.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# SQLite
# ─────────────────────────────────────────────────────────────────────────────

SCHEMA = """
PRAGMA foreign_keys = ON;

DROP TABLE IF EXISTS gear_items;
DROP TABLE IF EXISTS businesses;

CREATE TABLE businesses (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    slug            TEXT NOT NULL UNIQUE,
    name            TEXT NOT NULL,
    business_type   TEXT NOT NULL,
    address         TEXT,
    city            TEXT NOT NULL,
    state           TEXT NOT NULL,
    zip             TEXT,
    phone           TEXT,
    email           TEXT,
    website         TEXT,
    hours           TEXT,
    season          TEXT,
    notes           TEXT,
    source          TEXT
);

CREATE TABLE gear_items (
    id               INTEGER PRIMARY KEY AUTOINCREMENT,
    business_id      INTEGER NOT NULL REFERENCES businesses(id) ON DELETE CASCADE,
    category         TEXT NOT NULL,
    name             TEXT NOT NULL,
    brand            TEXT,
    description      TEXT,
    sizes            TEXT,
    price            REAL NOT NULL,
    price_unit       TEXT NOT NULL,
    price_weekend    REAL,
    price_peak       REAL,
    price_full_day   REAL,
    image_url        TEXT NOT NULL DEFAULT ''
);

CREATE INDEX idx_gear_business ON gear_items(business_id);
CREATE INDEX idx_gear_category ON gear_items(category);
"""


def build_sqlite() -> tuple[int, int]:
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)

    conn = sqlite3.connect(DB_PATH)
    try:
        conn.executescript(SCHEMA)

        biz_count = 0
        gear_count = 0
        for biz in BUSINESSES:
            cur = conn.execute(
                """
                INSERT INTO businesses
                    (slug, name, business_type, address, city, state, zip,
                     phone, email, website, hours, season, notes, source)
                VALUES (:slug, :name, :business_type, :address, :city, :state,
                        :zip, :phone, :email, :website, :hours, :season, :notes, :source)
                """,
                {**biz, **{k: biz.get(k) for k in
                           ("address", "zip", "phone", "email", "website",
                            "hours", "season", "notes", "source")}},
            )
            biz_id = cur.lastrowid
            biz_count += 1

            for item in biz["gear"]:
                conn.execute(
                    """
                    INSERT INTO gear_items
                        (business_id, category, name, brand, description, sizes,
                         price, price_unit, price_weekend, price_peak, price_full_day,
                         image_url)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        biz_id,
                        item["category"],
                        item["name"],
                        item.get("brand", ""),
                        item.get("description", ""),
                        item.get("sizes", ""),
                        float(item["price"]),
                        item["price_unit"],
                        item.get("price_weekend"),
                        item.get("price_peak"),
                        item.get("price_full_day"),
                        item.get("image_url", ""),
                    ),
                )
                gear_count += 1

        conn.commit()
        return biz_count, gear_count
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")       # deep blue
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")    # mid blue
ZEBRA_FILL = PatternFill("solid", fgColor="F2F2F2")        # light gray
ACCENT_FILL = PatternFill("solid", fgColor="E7F3FF")       # very light blue
TOTAL_FILL = PatternFill("solid", fgColor="FFD966")        # gold

HEADER_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=20, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="595959")
BODY_FONT = Font(name="Calibri", size=11, color="262626")
CATEGORY_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")
BOLD = Font(name="Calibri", size=11, bold=True, color="262626")

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def _fmt_price(item: dict) -> str:
    base = f"${item['price']:,.2f} / {item['price_unit']}"
    extras = []
    if item.get("price_weekend") is not None:
        extras.append(f"${item['price_weekend']:,.2f} weekend")
    if item.get("price_peak") is not None:
        extras.append(f"${item['price_peak']:,.2f} peak")
    if item.get("price_full_day") is not None:
        extras.append(f"${item['price_full_day']:,.2f} full day")
    if extras:
        base += "   (" + " • ".join(extras) + ")"
    return base


def build_excel() -> None:
    wb = Workbook()

    # ── Sheet 1: Overview ──────────────────────────────────────────────────
    overview = wb.active
    overview.title = "Overview"
    overview.sheet_view.showGridLines = False
    _build_overview(overview)

    # ── Sheet 2: All Gear (flat master list) ──────────────────────────────
    all_gear = wb.create_sheet("All Gear")
    all_gear.sheet_view.showGridLines = False
    _build_all_gear(all_gear)

    # ── Sheet 3+: One sheet per business ──────────────────────────────────
    for biz in BUSINESSES:
        ws = wb.create_sheet(_sheet_title(biz["name"]))
        ws.sheet_view.showGridLines = False
        _build_business_sheet(ws, biz)

    # ── Final sheet: Price Cheat Sheet ─────────────────────────────────────
    cheat = wb.create_sheet("Price Cheat Sheet")
    cheat.sheet_view.showGridLines = False
    _build_cheat_sheet(cheat)

    wb.save(XLSX_PATH)


def _sheet_title(name: str) -> str:
    # Excel sheet names are limited to 31 chars and cannot contain []:*?/\
    safe = "".join(c for c in name if c not in "[]:*?/\\")
    return safe[:31]


def _set_row_height(ws, row: int, height: float) -> None:
    ws.row_dimensions[row].height = height


# ── Overview sheet ────────────────────────────────────────────────────────────
def _build_overview(ws) -> None:
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 26

    ws.merge_cells("B2:H2")
    ws["B2"] = "Incline Village, NV — Outdoor Adventure Rental Directory"
    ws["B2"].font = TITLE_FONT
    ws["B2"].alignment = LEFT
    _set_row_height(ws, 2, 32)

    ws.merge_cells("B3:H3")
    ws["B3"] = (
        f"Compiled {datetime.now().strftime('%B %Y')} from each business's "
        "official website. Covers rental catalogs for skis/snowboards, mountain "
        "& electric bikes, paddleboards, kayaks, snowshoes and more."
    )
    ws["B3"].font = SUBTITLE_FONT
    ws["B3"].alignment = LEFT_TOP
    _set_row_height(ws, 3, 28)

    headers = [
        "#", "Business", "Type", "Address", "Phone", "Website", "Season", "Gear Items",
    ]
    row = 5
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    _set_row_height(ws, row, 26)

    for i, biz in enumerate(BUSINESSES, start=1):
        r = row + i
        values = [
            i,
            biz["name"],
            biz["business_type"],
            f'{biz.get("address","")}, {biz["city"]}, {biz["state"]} {biz.get("zip","")}'.strip(", "),
            biz.get("phone", ""),
            biz.get("website", ""),
            biz.get("season", ""),
            len(biz["gear"]),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = LEFT_TOP if col_idx in (3, 4, 7) else LEFT
            cell.border = BORDER
            if i % 2 == 0:
                cell.fill = ZEBRA_FILL
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=8).alignment = CENTER
        if biz.get("website"):
            ws.cell(row=r, column=6).hyperlink = biz["website"]
            ws.cell(row=r, column=6).font = Font(
                name="Calibri", size=11, color="1F4E79", underline="single"
            )
        _set_row_height(ws, r, 34)

    # Totals strip
    last_row = row + len(BUSINESSES) + 2
    total_items = sum(len(b["gear"]) for b in BUSINESSES)
    ws.cell(row=last_row, column=2, value=f"{len(BUSINESSES)} businesses").font = BOLD
    ws.cell(row=last_row, column=2).fill = TOTAL_FILL
    ws.cell(row=last_row, column=8, value=f"{total_items} total items").font = BOLD
    ws.cell(row=last_row, column=8).fill = TOTAL_FILL
    ws.cell(row=last_row, column=8).alignment = CENTER

    ws.freeze_panes = "A6"


# ── Master gear sheet ─────────────────────────────────────────────────────────
def _build_all_gear(ws) -> None:
    widths = [4, 24, 20, 16, 30, 18, 18, 22, 48]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A2:I2")
    ws["A2"] = "All Rental Gear — Master List"
    ws["A2"].font = TITLE_FONT
    ws["A2"].alignment = LEFT
    _set_row_height(ws, 2, 30)

    headers = ["#", "Business", "Category", "Brand", "Item", "Sizes", "Price", "Notes", "Image URL"]
    header_row = 4
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    _set_row_height(ws, header_row, 26)

    r = header_row + 1
    idx = 0
    for biz in BUSINESSES:
        for item in biz["gear"]:
            idx += 1
            values = [
                idx,
                biz["name"],
                item["category"],
                item.get("brand", ""),
                item["name"],
                item.get("sizes", ""),
                _fmt_price(item),
                (item.get("description") or "")[:200],
                item.get("image_url", ""),
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=col_idx, value=val)
                cell.font = BODY_FONT
                cell.alignment = LEFT_TOP if col_idx in (5, 8, 9) else LEFT
                cell.border = BORDER
                if col_idx == 9 and val:
                    cell.hyperlink = str(val)
                    cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
                if idx % 2 == 0:
                    cell.fill = ZEBRA_FILL
            ws.cell(row=r, column=1).alignment = CENTER
            _set_row_height(ws, r, 36)
            r += 1

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:I{r - 1}"


# ── Per-business sheet ────────────────────────────────────────────────────────
def _build_business_sheet(ws, biz: dict) -> None:
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 28
    ws.column_dimensions["H"].width = 44

    # Title block
    ws.merge_cells("A2:H2")
    ws["A2"] = biz["name"]
    ws["A2"].font = TITLE_FONT
    ws["A2"].alignment = LEFT
    _set_row_height(ws, 2, 32)

    ws.merge_cells("A3:H3")
    ws["A3"] = biz["business_type"]
    ws["A3"].font = SUBTITLE_FONT
    ws["A3"].alignment = LEFT
    _set_row_height(ws, 3, 20)

    # Info strip
    info = [
        ("Address", f'{biz.get("address","")}, {biz["city"]}, {biz["state"]} {biz.get("zip","")}'.strip(", ")),
        ("Phone", biz.get("phone", "")),
        ("Website", biz.get("website", "")),
        ("Season", biz.get("season", "")),
        ("Hours", biz.get("hours", "")),
        ("Notes", biz.get("notes", "")),
    ]
    r = 5
    for label, val in info:
        ws.cell(row=r, column=1, value=label).font = BOLD
        ws.cell(row=r, column=1).fill = ACCENT_FILL
        ws.cell(row=r, column=1).alignment = LEFT
        ws.cell(row=r, column=1).border = BORDER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        c = ws.cell(row=r, column=2, value=val)
        c.font = BODY_FONT
        c.alignment = LEFT_TOP
        c.border = BORDER
        if label == "Website" and biz.get("website"):
            c.hyperlink = biz["website"]
            c.font = Font(name="Calibri", size=11, color="1F4E79", underline="single")
        _set_row_height(ws, r, 28 if label != "Notes" else 46)
        r += 1

    r += 1  # spacer

    # Gear table
    headers = ["#", "Category", "Item", "Brand", "Sizes", "Price", "Description", "Image URL"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    _set_row_height(ws, r, 26)
    header_row = r
    r += 1

    # Group by category with subheader rows
    current_cat = None
    idx = 0
    for item in sorted(biz["gear"], key=lambda x: (x["category"], x["name"])):
        if item["category"] != current_cat:
            current_cat = item["category"]
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            c = ws.cell(row=r, column=1, value=current_cat)
            c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            c.fill = SUBHEADER_FILL
            c.alignment = LEFT
            c.border = BORDER
            _set_row_height(ws, r, 22)
            r += 1

        idx += 1
        img = item.get("image_url", "") or ""
        values = [
            idx,
            item["category"],
            item["name"],
            item.get("brand", ""),
            item.get("sizes", ""),
            _fmt_price(item),
            item.get("description", ""),
            img,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = LEFT_TOP
            cell.border = BORDER
            if col_idx == 8 and img:
                cell.hyperlink = img
                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
            if idx % 2 == 0:
                cell.fill = ZEBRA_FILL
        ws.cell(row=r, column=1).alignment = CENTER
        _set_row_height(ws, r, 42)
        r += 1

    ws.freeze_panes = f"A{header_row + 1}"


# ── Price cheat sheet ─────────────────────────────────────────────────────────
def _build_cheat_sheet(ws) -> None:
    """Quick, side-by-side price comparison for common rental categories."""

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 22

    ws.merge_cells("A2:E2")
    ws["A2"] = "Price Cheat Sheet — Incline Village Rentals"
    ws["A2"].font = TITLE_FONT
    ws["A2"].alignment = LEFT
    _set_row_height(ws, 2, 30)

    ws.merge_cells("A3:E3")
    ws["A3"] = "Cheapest daily price in each common category across Incline Village shops."
    ws["A3"].font = SUBTITLE_FONT
    ws["A3"].alignment = LEFT

    headers = ["#", "Category Bucket", "Cheapest Business", "Item", "Daily Price"]
    r = 5
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    _set_row_height(ws, r, 26)
    r += 1

    # Category buckets: (bucket name, keyword matchers for gear category)
    buckets = [
        ("Adult Ski Package", ["ski package", "skis"]),
        ("Adult Snowboard Package", ["snowboard"]),
        ("Cross-Country Skis", ["cross"]),
        ("Snowshoes", ["snowshoe"]),
        ("Hardtail Mountain Bike", ["front suspension", "hardtail"]),
        ("Full-Suspension MTB", ["full suspension"]),
        ("E-MTB", ["e-mtb", "e mountain", "electric mountain", "pedal assist"]),
        ("Cruiser / City Bike", ["cruiser"]),
        ("Kids Bike", ["kid"]),
        ("Stand-Up Paddleboard", ["paddleboard"]),
        ("Kayak (Day Rental)", ["kayak"]),
    ]

    idx = 0
    for bucket_name, needles in buckets:
        # Find all daily-rate items matching this bucket
        matches = []
        for biz in BUSINESSES:
            for item in biz["gear"]:
                blob = (item["category"] + " " + item["name"]).lower()
                if not any(n in blob for n in needles):
                    continue
                unit = item["price_unit"].lower()
                # Include only items whose price is expressed per day (or convertible)
                if "day" in unit or "24hr" in unit or "full day" in unit:
                    price = item["price"]
                elif item.get("price_full_day") is not None:
                    price = item["price_full_day"]
                else:
                    continue
                matches.append((price, biz["name"], item["name"]))
        if not matches:
            continue
        matches.sort(key=lambda x: x[0])
        price, biz_name, item_name = matches[0]

        idx += 1
        values = [idx, bucket_name, biz_name, item_name, f"${price:,.2f}/day"]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = LEFT_TOP
            cell.border = BORDER
            if idx % 2 == 0:
                cell.fill = ZEBRA_FILL
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=5).alignment = RIGHT
        ws.cell(row=r, column=5).font = BOLD
        _set_row_height(ws, r, 28)
        r += 1

    ws.freeze_panes = "A6"


# ─────────────────────────────────────────────────────────────────────────────
# Entry-point
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    biz_count, gear_count = build_sqlite()
    build_excel()
    print(f"✓ SQLite database written: {DB_PATH}")
    print(f"  {biz_count} businesses, {gear_count} gear items.")
    print(f"✓ Excel spreadsheet written: {XLSX_PATH}")


if __name__ == "__main__":
    main()
